import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

st.set_page_config(page_title="Check Payins", page_icon="📊", layout="wide")

st.markdown("""
<style>
    .stApp { background-color: #05051a; }
    .main { background-color: #05051a; }
    .block-container { padding: 1rem 2rem 2rem; }
    section[data-testid="stSidebar"] { background-color: #07071f; border-right: 2px solid #1a1aff; }
    section[data-testid="stSidebar"] * { color: #a0b4ff !important; }
    section[data-testid="stSidebar"] input { background-color: #0a0a2e !important; border: 1px solid #1a1aff !important; color: #ffffff !important; border-radius: 6px !important; }
    h2, h3 { color: #a0b4ff !important; }
    p, span, label, div { color: #c8d4ff; }
    div[data-testid="metric-container"] { background: #0a0a2e; border: 1px solid #1a1aff; border-radius: 10px; padding: 1rem; }
    div[data-testid="metric-container"] label { color: #7a94ff !important; }
    div[data-testid="metric-container"] div { color: #ffffff !important; }
    .stRadio label { color: #a0b4ff !important; }
    input[type="number"] { background-color: #0a0a2e !important; color: #ffffff !important; border: 1px solid #1a1aff !important; }
    .stDownloadButton button { background-color: #1a1aff !important; color: #ffffff !important; border: none !important; border-radius: 8px !important; font-weight: 700 !important; padding: 0.6rem 1.5rem !important; }
    .stDownloadButton button:hover { background-color: #3333ff !important; }
    .stAlert { border-radius: 8px; background-color: #0a0a2e !important; border-left: 4px solid #1a1aff !important; color: #c8d4ff !important; }
    .stDataFrame { border: 1px solid #1a1aff; border-radius: 8px; }
    .stFileUploader { border: 1px dashed #1a1aff !important; border-radius: 8px; background-color: #0a0a2e !important; }
    .streamlit-expanderHeader { color: #7a94ff !important; background-color: #0a0a2e !important; border: 1px solid #1a1aff !important; border-radius: 8px !important; }
    .streamlit-expanderContent { background-color: #0a0a2e !important; border: 1px solid #1a1aff !important; }
    hr { border-color: #1a1aff !important; }
    .stRadio div { background: transparent !important; }
    .stSelectbox > div { background-color: #0a0a2e !important; }
    .rule-badge { display: inline-block; background: #0a0a2e; border: 1px solid #1a1aff; border-radius: 20px; padding: 3px 10px; margin: 3px; font-size: 12px; color: #a0b4ff; }
    .proc-title { color: #ffffff; font-weight: 700; font-size: 13px; margin-bottom: 4px; }
</style>
""", unsafe_allow_html=True)

# ── REGLAS HARDCODEADAS ────────────────────────────────────────────────────────
# Cada entrada: (fragmento a buscar en Complementary info, Procesador)
# La búsqueda es case-insensitive y por substring
RULES = [
    ("SERVIPAG CASH COLLECTION",     "SERVIPAG"),
    ("0780537906",                    "SERVIPAG"),
    ("ABONO TBK TC",                  "TRANSBK"),
    ("ABONO TBK TD",                  "TRANSBK"),
    ("ABONO VENTAS GETNET",           "GETNET"),
    ("TRANSF ALTOMON 995469006",      "KLAP"),
    ("TRANSF ALTOMONT DE 995469006",  "KLAP"),
    ("TRANSF ALTOMONT DE 775979844",  "KUSHKI"),
    ("TRANSFERENCIA DE KUSHKI CHILE", "KUSHKI"),
    ("Pago Proveedores 995469006",    "MERCADO PAGO"),
    ("Pago Proveedores 714009001",    "MERCADO PAGO"),
    ("Pago Proveedores 772140665",    "MERCADO PAGO"),
    ("Pago Proveedor 76516950",       "MERCADO PAGO"),
    ("Pago Proveedor 77214066",       "MERCADO PAGO"),
    ("Pago Proveedores 769237836",    "MERCADO PAGO"),
    ("Transferencia De D-local Chile","MERCADO PAGO"),
    ("TRANSFERENCIA DE",              "KHIPU"),
    ("TRANSF. DE",                    "KHIPU"),
]

def get_processor(concept):
    if not isinstance(concept, str): return None
    for keyword, processor in RULES:
        if keyword.upper() in concept.upper():
            return processor
    return None

# ── HEADER ─────────────────────────────────────────────────────────────────────
logo_b64 = "UklGRmgTAABXRUJQVlA4WAoAAAAQAAAAJQIAJQIAQUxQSLwPAAAB8Ef+3zfF///dHU6nMcYYY4yMMcZIkqwkK2slSZKVJEmykpW1krWykrVkrZWVx48kayXJykpWstZKkqwkSZIkSTKSjDHGaZxOh/sfu3ue9+M4znP2+fp5RsQEgO9/3/++/33/+/73/e/73/e/73/f/77/ff/7/vf97/vf97/vf9//vv99//v+9/3v+9/3v+9/3/++/33/+09mpK7GaP/DwUi9p3L9/ML54UjAveIrJjo3lyP/yWCzJhJuJr1Gm0XSbK17dZaQstD2n4yMhaTDXqMv0dzVudcgkhqP/pNRjbTjvv98//n+8/33/xWI1dxv6ejt62tvaayNMVdhFfeau/vGXr169byv62FNTKVAvKa+uaWjt6+vt7WlqTYdKWtYOF3X/eLV4uLi4qux3vuZKCv79GTD8NzW0dVtwUBEo3B3fbQ9P9yQ1F2ARWv6/9o8yRYs/LWVuzpef9OeCcrGYlUPn06v7R1fZfMFAxF5sZC/Pjvcnht9WBlmZYgW6prZuOBo92rn4+OUXsZpiYGlYxMpS6efhyt0pVjw4fR2Hkn55dpwhSaNFqx5urB7y5GaX29MtQZZeRHs/pRHUnN7rLpMi/R/LqFIY30wqk795CkKLX7u1KRIDi5cchR/NdOslQ/psdMSkvP8WndF2cXSE+cmijavJmuYCoHW1RxH0aX9oZio9MDanYmSFvf6w+VB/MUJR7HG7nGZlR4/4Sgjv5jKSBdoXs2jlOZ2V0BAsGHm1ESZS1u9Ie/TWvc4Klq2BPpOOEp78SQkV8VsHqUtLWQYkd72uYjSl5ZTXhebMVHZcqVmBeXeaJRI771Cqa8HNZL6DQuVvO3RvIw1Hlj4Z6E/ueKS4e2oLkti3kDJjbkoQfQSVS2+DXmX3plFlcuSyHsD5TdnKqRg93YslJ5/yTiLl5TB0nzYq9iTPP5hJJZRzfWUDE0XqORW2lFgRR20/tI96nkR/zCSW6jqcaW4gVtUdC/tBBpK6iCOaV7EnhXwDyOxzZXBw2pRgzlUle8nnbAVlfLdHsQ6CviHEf/OUeHtSiGsKYcKTzEH0JJTCA/T3lN3iX8YwRlUeyUiovYCVZ7TnGifVeJTnhPfxT+NCa4YnxOQ2EWVD9LguO5WISzWegybsf402m9RdaOHUekfuULWagqc67Mq4QfNW1oL+IcROUT1L9JUvUVU15yLA2VtgYAb+btfFk1RuZSnhPbxD0Ob5C6AszpN+ArV5eNBINXn7XHjYOn1UNfD+79sH5hYzXIR+MxL2HP806gromgje319y0UV20m0aRR+fbC7u3teosgNAXV14Tf8bHmkTgfnNeO7nG4j7CGZUyHm0afRlkwmk7nX9Wrl2CiPQp9Q6PX6eHM6kUgka/pmD0si8HuQoj4nhJ8tP6lNRKPRaEWmb+mMOzjvYGTs40/m8XRTPAC0LNq5a1Hd1HjIOBeQW+1OMPg9S7VOn5XKoJaCAOvsZa0ONhNd34oCip0E2kcUWDoYrmRgU6scueR2DhoZ0Nef89zX3hQIrXhfJMJe70icIP12kwaOo4/3yp9FpC9OVoBT/dEZHW4R1GUFXD8Lg+PUgvm7vRQIvf/0AQPR+nOLaNo7nllk+RdAq3UcqRJ4/41yKeM+9+/I+G4VUEY+mGSlJmfTSG59jgMlG8j/wlqKgPraBKfZ8Ay2i9SXPRoRQHNJkYclJB1zn0lOZS2ngDb4vEiFs5qT+CWZ8SYEtFpHARGLbyLghsFTmuOIV9QaVLlWBuQZQ5EWk2bKdbQzpF6OAbX+0qS6jjnpK1GVXgWBmj0rYH5IB3cctUiuqj2CTXIiPsSgTGu0qA4SQK8tcSLscMCWkfqvANBrA9Mt4Ja1NyQ39R4R3EZaPg8iyxv2DonvakFkbJ9qXrMXylNt6uCN2ilJoc0jUgbRWaZsC28Tma9BbHeBaL/CXjMS390HVw2n7jW2/vphfWVcswPfSIpe0Ye0/CWUbckC0XFCUGiNKF9v7zURnwbX1GP1owvbx9kCx18ad+f7X//pr45qv1rzshmi88ryrRWJR0F0OxEO2Ap+IbqqcQutY/6EI6l5ONuue114k2gGyrcxootKYZF9ohlbiROiWXDH6leXHAXy66l6yJx7WOqEhj8o4z4RfdKFsUmiTVvVJRqr2RVSb645Cua3m+fcw+7laC5i5Vtok2gAxDdZNMdhO61Ie5JwgUDPIarrFc1IuxIo3xKHRBkJQkWay4ydp0TLunrh2RKWPY+JJqF8q7qgOQtLAHs0N/V2XhM9B+UzOxzLn3EaPljG1d/RbARlWKYptNqZI3qkXPU2qu0Vb2ms1nIuR7OoyzBFU2y3M09Uo1p6C8s3s6aMa7JoppgML2hKnXY2iSKKaV94ebRMlCnjWpD2Dcj4hAYfiwuopb2ysDz6/B+PQZery2OZ9PHPg9O8lWKUxuyxs0EUUiq4jOXSWxqroYxryNMs6jK8pym221kiqlWq1RBinG5tn9+WN+0u1Uz02mXq72jWAzIsiftA1KoSW0Fya+v1g0hQDwRD8QfDiwcl73tOw4dFVakS+X5MudvgMtWXNMdhGfZpbhvsvCMaVak2S1Xa7IiAXVbRMHlkeNwADb4T1WQqAqE4ZRhcJnVMU4pLECjQXFXbGSH6yBQa4UTFsQg4j/X9MD2tiWg9KGjAUkVZuUJbNPhIgiaL5jRqp4NoN6bQV6Qt9TIg1fuvvOxenuYiJoa9xzIGlonmmDD2Fml3wG6NSWPUqxPM0ZReMqDuK3lY+pSGPxQT2iprxokOE8IiB0TztpIXNPiOKVNv0uxEgPz+nYdFt2lwWkxVyWsey9VOZD4S1sqJntkKfyc6rFCm3yLhg0Df4GXwkegoKWQcvabDkCpRoMGvuqDgF6QtPbDFpomsp8pMIOl1dZnyjIgPikifeE57kcQaoAptEhV7BfUYRMdJW9BJhJdJRdhfND+iAroML6vnNLgXp2PjnCjf4DH4lIq9IcKjiJDEPhKv6PZiJSL+QVdD/0TzRRMwweVJE80pFbkisiboai+Rutm92ojeUkGLRcSndAHaB041DPbZVyI0e5gA9miikUm1AvTsB8rDiFaU0peIsNRIFdpC72m4oVkn0y6I0BpiZNooR+Ji0gEMW0R40ySg38B8D5NpK0LXYMlk0ewFVYIBiwjP79OE5rgH1VzTXOtU8IYTYW5IIwo8NZB6NeAknaXCizayxgIiFseCEp1lyKIbKNMlTbZOqXiOCo8eMIKKGQs9KH5GYzWS1d9SYW48RBKZMpDa7AHHi2R41a+TaI/O8GdjNkqyQGO2UwXeWFLt01iDSsEKGV70605YzTeOAh+5l3ZAg3M6FVslQ/4pwxxpNd8tJL8IOWstkGHpXZw500fv8NfWWgUBvKPBZZ1GGzdQqhUa3GBKdRlkiGutmq3k6yIKHXEvWCUqNFJBU4EM8XYy46B2poD0fAycB9fpEM9HIg5Y4wZHm/v3CEY5TbGfJPTBRKngDVGpR6nIrgAsHky0Vsc1CFQ09C/ecBQ772KvifBHFVVgSQDy3Kf+e7FfxOv713IcBZ4kCaDFEIDWxeum2G9Ypme1gPYvOpmjNpMG71qZo2DnPkdy3k/z2KTB7AOmEDznAhCxeHn0Y2fv+MZC4ccR9+qi4vttIRqoywtAROv6aGdp4dPO0Y2FQq2nQBlcFIGIucPVyaGenuF3X89NdJx9qjuJF4nw5mnMFqvoWs+jyEGaB3kivOiPKpQ6FSOx8di97t0RIZa+jjbGg8FEXVfSnjbFhUi7EyOByoIYwdZ40AHsU6H1YzQR0BjT9EC8a/YABb+iiV5RYWlrvCkZDMbq2quYdDDiEphr19wqukNm96TSFoSPXcBoBOJxSx3ExbiDcU6FiNbR2sePn7YvUcJNGlgjs1u8L19k3yWw+G28O+VK7C8JeLs9aL1Tjr/VqELrKvHtGnv1twJkLjbSjHBxOCofdBku8fNuzI2gjYvDTgfaS1O17QiQ158qhPys2pa26gp4WKdT1N1I8FYBbUk+bpBh2pXC5xKMOIDwgmLndSCws6AQ4oJuBzoMV8DrhWfdFY6CX10JUofS3czTVboSe8XFjTqByDpXKdsKItljS6VpZivwyR1+vqpyAr0SzKoAbXeSmS/e0LW4EqRuxI05guSeQoU+JgS00YIyfDsG9muyroEPHcX3xS0roY+W5NqOP+Nkre4E77mwv5xB9SZXJTfEQLD+vKAI30yBQzZiSsY/fJGGDXFXAu0vS6ZcC3SaHpM+FTZLAImvitz0gXh9xFBjKw2OtUUu12ag2aJqdwSBQ2FruhKgzVnyWE8YJIoeAz2GAhCYslQ4agQZWWtWhYUYEMa+SLVfAdWXVEPO4MGdqM2wGhCeNmSx/gkAsA2vCUxxQV9JQH98KZ25kmJSANR9t2TLTQSBNLEi0V4dQHBdHvbccCcIvCjIYc6FAACGTI+ByBcu5jsNsPvrJbmuX4ZB2vjrnFTWQYcOxBXLpiT8ew0AQD+XBoJTppjduCrAuq9lsGbD8HPsh9dA7LsKAIGnWYn4lwaQmTXucnlK75NAH5wypDA/ROHn8J48EJzmQo5TygAkF0rCcoM6/Lrxxmsg+M4QcUYGEHtzI4n1o4OB5FrfniSlL3UgtnlXgr1O+G1jTh6AsbxLQbDjhynE3HzA4Les7cZrQO885HQXAkCrfHNmiTM2+qIgP4sN7JbE5VZaQiA6NnbOhVgno3H4PevOSqQ1bVp0F5UqAcQe7xpk1vGzONhlTYcmRYubASQnz0yqcxEAUDm8fcdFmFeLbSFQNPZoKWuJMM7/eaCBjFWT5yZZ/sfTNNhm7QcWwWMigNjQsUGVrVELINS5YpDwo2dRcBp7ce7sMuVuAPGBjRIJ/0sQAKsf+2FS5Zb7U6AySw+vFqiy811RkDY2sGZQ8KP3D3VwXPHi0tHVPTKASO+nIs1SUDUACHf9s33L7fCbjcl6BpTs/uSXw6yJiMXro+/zEx0R8MBQ++TyzlnW+imfPdn9MjfRk2HCfg61T67u33E75tXOwki9Di6oNb5c3LnidszL7Y/DtQwkD7e9XT/K/S67v/a+pwKItcbXK/tX2Ww2e733bX6iMwJitYfji1un2dJPhez57tr8q957DFyRRVJ1PaN/zc7Ozs5MDDYkw0AfjCUrq6qq0sl4WAPP1CMVycqqqqqqdDIeDTKQOBBL1bT2j/78rLO5MhHRwDX1SKKypXv055GupkwiooGSwXi6qq2jo6OjoTIZC4LQQCyR/Dka0kBGPVKRzFRVVVWlkxXRoAa+/33/+/73/e/73/e/73/f/77/ff/7/vf97/vf97/vf9//vv99//v+9/3v+9/3v+9/3/++/33/+/73/e/7/38jBVZQOCCGAwAAkFYAnQEqJgImAj4xGItEoiGhEB1UBCADBLS3cLr3AH8A9uHQH4qov+AP4B+AGqFcA/AD9gKr/2gP4B+QGKAfwD2/7sA/wH8A7P/Ux/AH8P7APb+N+ADflgHvtk5D32ych77ZOQ99snIe+2TkPfbJyHvtk5D32ych77ZOQ99snIe+2TkPfbJyHvtk5D32ych77ZOQ99snIe+2TkPfbJyHvtk5D32ych77ZOQ99snIe+2TkPfbJyHvtk5D32ych77ZOQ99snIe+2TkPfbJyHvtk5D32ych77ZOQ99snIe+2TkPfbJyHvtk5D32ych77ZOQ99snIe+2TkPfbJyHvtk5D32ych77ZOQ99snIe+2TkPfbJyHvtk5D32ych77tjEH5KtzjshteLk5D322J6hmPuTkTbjwh6HwPfbJyHvuJ7N95XISdwJhtplMfm6VOzyHfJLf+ydvsRBre2hH+NLqUK5AE216zKfZdQDBpwhzrXeU5bTln18GAFo8xJu5sGAPyHcG+1SzXcXylQ4ZvQkkj3xC0Q94WGjkV2g5ljYVicSmVJkjVCBWUHtp291csW2MFKTqfgzWGojS4Vq7kxWllqGwOPHoVP0MZK6FaNSKITb2i8/944MaB549B+B4Zef5ylF4yB3WPamSIR49tWZo5lqPYhknpQXymbXi5OQ99snIe+2TkPfbJyHvtk5D32ych77ZOQ99snIe+2TkPfbJyHvtk5D32ych77ZOQ99snIe+2TkPfbJyHvtk5D32ych77ZOQ99snIe+2TkPfbJyHvtk5D32ych77ZOQ99snIe+2TkPfbJyHvtk5D32ych77ZOQ99snIe+2TkPfbJyHvtk5D32ych77ZOQ99snIe+2TkPfbJyHvtk5D32ych77ZOQ99snIe+2TkPfbJyHvtk5D32ych77ZOQ99snIe9IAA/vTrIAAAAAAAAAAAAAAAVHO81nO8l9zvIBGvMMvTbiD/Y6jzKvj6uavj6vAoNp8df81aGBZGBvNR6L75ADNPBx2ewpCqaKH/qdNFvP+jpalOBVWV/vOjt0ec1P/1MqB6loc+8nerQaAEA9o2n7nYP/QYXPE6QUq9p5HCeIQ6YHlfdcTn0QE8ho2WI+Li3SnW0Ou0tthc8NsPUuKIjAkXr3i5UytvykjAePPBjRMuvMLYgDjXS/AiUQAAAAAAAAAAAAAAAAAAAAA="
st.markdown(f"""
<div style="background: #0a0a2e; padding: 2rem 2.5rem; border-radius: 14px; margin-bottom: 1.8rem; display: flex; align-items: center; justify-content: center; gap: 2.5rem; border-bottom: 3px solid #1a1aff; min-height: 110px;">
    <img src="data:image/webp;base64,{logo_b64}" style="height: 80px; filter: brightness(0) invert(1); display: block;">
    <div style="border-left: 2px solid #1a1aff; padding-left: 2rem; display: flex; align-items: center;">
        <div style="color: #ffffff; font-size: 2.2rem; font-weight: 800; letter-spacing: -1px; line-height: 1;">Check Payins</div>
    </div>
</div>
""", unsafe_allow_html=True)

# ── PANEL DE REGLAS ────────────────────────────────────────────────────────────
rules_by_proc = {}
for keyword, processor in RULES:
    rules_by_proc.setdefault(processor, []).append(keyword)

with st.expander("📋 Reglas de identificación cargadas"):
    cols = st.columns(4)
    for i, (proc, keywords) in enumerate(sorted(rules_by_proc.items())):
        with cols[i % 4]:
            badges = "".join(f'<span class="rule-badge">{k}</span>' for k in keywords)
            st.markdown(f'<div class="proc-title">{proc}</div>{badges}', unsafe_allow_html=True)

st.markdown("")

# ── SIDEBAR ────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.header("⚙️ Configuración")
    st.markdown("**AR Processors — columnas**")
    col_date = st.text_input("Fecha", value="Payment date")
    col_amount = st.text_input("Monto", value="LC Amount")
    col_processor = st.text_input("Procesador", value="Processor")
    st.markdown("---")
    tolerance = st.number_input("Tolerancia sin alerta (%)", value=10, step=1, min_value=0, max_value=100)

# ── FILE UPLOADS ───────────────────────────────────────────────────────────────
col1, col2 = st.columns(2)
with col1:
    st.subheader("🏦 Extracto Bancario")
    banco_file = st.file_uploader("Excel del banco (Treasury Factory)", type=["xlsx","xls","csv"], key="banco")
with col2:
    st.subheader("📋 AR Processors")
    interno_file = st.file_uploader("Excel AR Processors", type=["xlsx","xls","csv"], key="interno")

# ── PARSE BANCO ────────────────────────────────────────────────────────────────
def parse_banco_excel(file):
    try:
        raw = pd.read_excel(file, header=None) if not file.name.endswith(".csv") else pd.read_csv(file, header=None)
        header_row = None
        for i, row in raw.iterrows():
            if any("Transaction date" in str(v) for v in row.values):
                header_row = i
                break
        if header_row is None:
            st.error("No encontré encabezados. Asegurate de usar el formato Treasury Factory.")
            return pd.DataFrame()

        df = pd.read_excel(file, header=header_row) if not file.name.endswith(".csv") else pd.read_csv(file, header=header_row)
        df.columns = ['Account code','Account ID','Transaction date','Value date',
                      'Description full','Description','Complementary info',
                      'Reference','Other reference','Debit','Credit']

        exclude = ['Opening balance','Closing balance','Description','Transaction date']
        df = df[df['Account code'].notna() & ~df['Description'].isin(exclude)].copy()
        df['Transaction date'] = pd.to_datetime(df['Transaction date'], errors='coerce')
        df = df[df['Transaction date'].notna()]
        df['Credit'] = pd.to_numeric(df['Credit'], errors='coerce').fillna(0)
        df['Debit']  = pd.to_numeric(df['Debit'],  errors='coerce').fillna(0)
        df['Processor'] = df['Complementary info'].apply(get_processor)
        return df
    except Exception as e:
        st.error(f"Error leyendo extracto: {e}")
        return pd.DataFrame()

# ── PARSE AR ───────────────────────────────────────────────────────────────────
def parse_interno(file, col_d, col_a, col_p):
    try:
        df = pd.read_csv(file) if file.name.endswith(".csv") else pd.read_excel(file)
        df.columns = [str(c).strip() for c in df.columns]
        missing = [c for c in [col_p, col_d, col_a] if c not in df.columns]
        if missing:
            st.error(f"Columnas no encontradas: {missing}. Disponibles: {list(df.columns)}")
            return pd.DataFrame()
        df = df[[col_p, col_d, col_a]].copy()
        df.columns = ['Processor','Date','Amount']
        df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
        df['Amount'] = pd.to_numeric(df['Amount'], errors='coerce').fillna(0)
        df = df[df['Date'].notna()]
        df['Day'] = df['Date'].dt.strftime("%d/%m")
        return df
    except Exception as e:
        st.error(f"Error leyendo AR Processors: {e}")
        return pd.DataFrame()

# ── BUILD EXCEL ────────────────────────────────────────────────────────────────
def build_excel(results, period):
    wb = Workbook()
    ws = wb.active
    ws.title = "Check Payins"
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr(r, c, v, bg="1F4E79", fg="FFFFFF"):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = Font(name='Arial', bold=True, color=fg, size=10)
        cell.fill = PatternFill('solid', start_color=bg)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    ws.merge_cells('A1:G1')
    t = ws.cell(row=1, column=1, value=f"Check Payins — {period}")
    t.font = Font(name='Arial', bold=True, size=13, color="1F4E79")
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 26

    headers = ['Procesador','Fecha','Banco (CLP)','AR Processors (CLP)','Diferencia (CLP)','Dif. %','Estado']
    for i, h in enumerate(headers, 1):
        hdr(2, i, h)
    ws.row_dimensions[2].height = 22

    r = 3
    for proc, rows in results.items():
        for idx, row in enumerate(rows):
            bg_alt = "F2F7FB" if idx % 2 == 0 else "FFFFFF"
            diff = row['diff']
            pct = abs(diff / row['interno'] * 100) if row['interno'] else None
            is_ok = pct is not None and pct <= 10
            if is_ok: st_bg, st_txt, st_lbl = "D9F7D9","1A6B1A","OK"
            elif diff < 0: st_bg, st_txt, st_lbl = "FFDAD9","8B1A1A","Cobro menor"
            else: st_bg, st_txt, st_lbl = "FFF3CD","7A5A00","Cobro mayor"

            for c in range(1, 8):
                cell = ws.cell(row=r, column=c)
                cell.border = border
                cell.font = Font(name='Arial', size=10)
                if c not in [5,6,7]: cell.fill = PatternFill('solid', start_color=bg_alt)

            ws.cell(row=r, column=1).value = proc
            ws.cell(row=r, column=2).value = row['fecha']
            ws.cell(row=r, column=2).alignment = Alignment(horizontal='center')
            for c, key in [(3,'banco'),(4,'interno')]:
                cell = ws.cell(row=r, column=c)
                cell.value = row[key] if row[key] is not None else 0
                cell.number_format = '#,##0;(#,##0);"-"'
                cell.alignment = Alignment(horizontal='right')
            diff_cell = ws.cell(row=r, column=5)
            diff_cell.value = f"=C{r}-D{r}"
            diff_cell.number_format = '#,##0;(#,##0);"-"'
            diff_cell.alignment = Alignment(horizontal='right')
            diff_cell.fill = PatternFill('solid', start_color=st_bg)
            pct_cell = ws.cell(row=r, column=6)
            pct_cell.value = f"=E{r}/D{r}" if row['interno'] else '-'
            pct_cell.number_format = '0.0%;(0.0%);"-"'
            pct_cell.alignment = Alignment(horizontal='center')
            pct_cell.fill = PatternFill('solid', start_color=st_bg)
            st_cell = ws.cell(row=r, column=7)
            st_cell.value = st_lbl
            st_cell.font = Font(name='Arial', bold=True, size=10, color=st_txt)
            st_cell.fill = PatternFill('solid', start_color=st_bg)
            st_cell.alignment = Alignment(horizontal='center')
            r += 1

        # Subtotal por procesador
        subtotal_start = r - len(rows)
        for c in range(1, 8):
            cell = ws.cell(row=r, column=c)
            cell.font = Font(name='Arial', bold=True, size=10, color="FFFFFF")
            cell.fill = PatternFill('solid', start_color="2d4a8a")
            cell.border = border
            cell.alignment = Alignment(horizontal='center' if c in [1,2,6,7] else 'right')
        ws.cell(row=r, column=1).value = f"TOTAL {proc}"
        ws.cell(row=r, column=3).value = f"=SUM(C{subtotal_start}:C{r-1})"
        ws.cell(row=r, column=3).number_format = '#,##0;(#,##0);"-"'
        ws.cell(row=r, column=4).value = f"=SUM(D{subtotal_start}:D{r-1})"
        ws.cell(row=r, column=4).number_format = '#,##0;(#,##0);"-"'
        ws.cell(row=r, column=5).value = f"=C{r}-D{r}"
        ws.cell(row=r, column=5).number_format = '#,##0;(#,##0);"-"'
        ws.cell(row=r, column=6).value = f"=E{r}/D{r}"
        ws.cell(row=r, column=6).number_format = '0.0%;(0.0%);"-"'
        r += 1

    ws.column_dimensions['A'].width = 16
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 16
    ws.freeze_panes = 'A3'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ── MAIN ───────────────────────────────────────────────────────────────────────
if banco_file and interno_file:
    with st.spinner("Procesando archivos..."):
        banco_df = parse_banco_excel(banco_file)
        ar_df = parse_interno(interno_file, col_date, col_amount, col_processor)

    if banco_df.empty or ar_df.empty:
        st.stop()

    # Procesadores identificados en el banco
    processors_in_banco = sorted(banco_df['Processor'].dropna().unique().tolist())
    processors_in_ar = sorted(ar_df['Processor'].dropna().unique().tolist())
    all_processors = sorted(set(processors_in_banco + processors_in_ar))

    if not all_processors:
        st.warning("No se identificaron procesadores. Revisá las reglas cargadas.")
        st.stop()

    st.markdown("---")
    st.subheader("⚙️ Selección")
    sel_col1, sel_col2 = st.columns([2,1])
    with sel_col1:
        selected_processors = st.multiselect(
            "Procesadores a analizar",
            options=all_processors,
            default=all_processors
        )
    with sel_col2:
        st.markdown("<br>", unsafe_allow_html=True)
        run = st.button("▶ Analizar conciliación", type="primary", use_container_width=True)

    if run and selected_processors:
        results = {}
        summary_rows = []

        for proc in selected_processors:
            # Banco por día
            banco_proc = banco_df[banco_df['Processor'] == proc].copy()
            banco_by_day = banco_proc.groupby(banco_proc['Transaction date'].dt.strftime("%d/%m"))['Credit'].sum().to_dict()

            # AR por día
            ar_proc = ar_df[ar_df['Processor'].str.upper() == proc.upper()].copy()
            ar_by_day = ar_proc.groupby('Day')['Amount'].sum().to_dict()

            all_dates = sorted(set(list(banco_by_day.keys()) + list(ar_by_day.keys())))
            if not all_dates:
                continue

            rows = []
            for d in all_dates:
                b = banco_by_day.get(d, 0) or 0
                i = ar_by_day.get(d, 0) or 0
                diff = b - i
                pct = abs(diff / i * 100) if i else None
                is_ok = pct is not None and pct <= tolerance
                rows.append({'fecha': d, 'banco': b, 'interno': i, 'diff': diff, 'pct': pct, 'is_ok': is_ok})

            results[proc] = rows

            total_b = sum(r['banco'] for r in rows)
            total_i = sum(r['interno'] for r in rows)
            diff_t = total_b - total_i
            pct_t = diff_t / total_i * 100 if total_i else 0
            dias_ok = sum(1 for r in rows if r['is_ok'])
            summary_rows.append({
                'Procesador': proc,
                'Banco (CLP)': total_b,
                'AR Processors (CLP)': total_i,
                'Diferencia (CLP)': diff_t,
                'Dif. %': f"{pct_t:+.1f}%",
                'Días OK': f"{dias_ok}/{len(rows)}",
                'Estado': "✅ OK" if abs(pct_t) <= tolerance else ("🔴 Cobro menor" if diff_t < 0 else "🟡 Cobro mayor")
            })

        # ── Resumen general
        st.markdown("---")
        st.subheader("📈 Resumen por procesador")
        st.dataframe(pd.DataFrame(summary_rows), use_container_width=True, hide_index=True)

        # ── Detalle por procesador
        st.markdown("---")
        st.subheader("📅 Detalle por día")

        selected_detail = st.selectbox("Ver detalle de:", list(results.keys()))
        filter_opt = st.radio("Mostrar:", ["Todos","Solo con diferencias","Solo OK"], horizontal=True)

        if selected_detail and selected_detail in results:
            detail_rows = []
            for r in results[selected_detail]:
                if filter_opt == "Solo con diferencias" and r['is_ok']: continue
                if filter_opt == "Solo OK" and not r['is_ok']: continue
                diff = r['diff']
                pct = r['pct']
                if r['is_ok']: estado = "✅ OK"
                elif diff < 0: estado = "🔴 Cobro menor"
                else: estado = "🟡 Cobro mayor"
                detail_rows.append({
                    "Fecha": r['fecha'],
                    "Banco (CLP)": f"{r['banco']:,.0f}",
                    "AR Processors (CLP)": f"{r['interno']:,.0f}",
                    "Diferencia (CLP)": f"{diff:+,.0f}",
                    "Dif. %": f"{pct:+.1f}%" if pct is not None else "-",
                    "Estado": estado
                })

            # Total row
            rows_sel = results[selected_detail]
            tb = sum(r['banco'] for r in rows_sel)
            ti = sum(r['interno'] for r in rows_sel)
            td = tb - ti
            tp = td / ti * 100 if ti else 0
            detail_rows.append({
                "Fecha": "**TOTAL**",
                "Banco (CLP)": f"**{tb:,.0f}**",
                "AR Processors (CLP)": f"**{ti:,.0f}**",
                "Diferencia (CLP)": f"**{td:+,.0f}**",
                "Dif. %": f"**{tp:+.1f}%**",
                "Estado": "✅ OK" if abs(tp) <= tolerance else ("🔴 Cobro menor" if td < 0 else "🟡 Cobro mayor")
            })

            st.dataframe(pd.DataFrame(detail_rows), use_container_width=True, hide_index=True)

        # ── Descarga
        st.markdown("---")
        all_dates_flat = []
        for rows in results.values():
            all_dates_flat.extend([r['fecha'] for r in rows])
        period = f"{min(all_dates_flat)} al {max(all_dates_flat)}" if all_dates_flat else ""
        excel_buf = build_excel(results, period)
        st.download_button(
            label="⬇️ Descargar Excel completo",
            data=excel_buf,
            file_name=f"check_payins_{datetime.now().strftime('%Y%m')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

else:
    st.info("👆 Cargá los dos archivos para comenzar el análisis.")
    with st.expander("ℹ️ ¿Cómo usar esta herramienta?"):
        st.markdown("""
        1. **Extracto bancario**: Excel exportado de Treasury Factory (todos los movimientos)
        2. **AR Processors**: Excel con pagos internos por procesador
        3. La herramienta identifica automáticamente cada procesador usando las reglas cargadas
        4. Seleccioná qué procesadores analizar y hacé clic en **Analizar**
        5. Revisá el resumen y el detalle por día, y descargá el Excel
        """)
